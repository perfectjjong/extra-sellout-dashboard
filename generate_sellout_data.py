"""
eXtra Sellout Dashboard - Data Generator
=========================================
주간 xlsx 데이터를 data.json으로 변환하여 대시보드 갱신.

Usage:
    python generate_sellout_data.py              # 기본 실행 (2026 데이터만 갱신)
    python generate_sellout_data.py --rebuild    # 전체 재빌드 (2024+2025+2026)

Input:
    - extra_2024.xlsx, extra_2025.xlsx (07.FCST_AI 폴더)
    - week01~52.xlsx (01. eXtra Raw/01. Sell out/00. Weekly 폴더)
    - item_master.json, promoter_config.json, branch_region.json

Output:
    - data.json (대시보드용 압축 데이터)
    - unclassified_items.json (미분류 신규 아이템 리포트)
"""

import os
import sys
import json
import re
import math
from datetime import datetime, timedelta
from collections import defaultdict

# ─── Paths ───
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FCST_DIR = r"C:\Users\J_park\Documents\2026\01. Work\07.FCST_AI"
WEEKLY_DIR = r"C:\Users\J_park\Documents\2026\01. Work\01. Sales\01. Sell out\01. Weekly\01. eXtra Raw\01. Sell out\00. Weekly"

ITEM_MASTER_PATH = os.path.join(SCRIPT_DIR, "item_master.json")
PROMOTER_CONFIG_PATH = os.path.join(SCRIPT_DIR, "promoter_config.json")
BRANCH_REGION_PATH = os.path.join(SCRIPT_DIR, "branch_region.json")
DATA_JSON_PATH = os.path.join(SCRIPT_DIR, "data.json")
UNCLASSIFIED_PATH = os.path.join(SCRIPT_DIR, "unclassified_items.json")

# ─── BTU → Ton Ranges ───
BTU_TO_TON = [
    (0, 16500, "1 Ton"),
    (16501, 20000, "1.5 Ton"),
    (20001, 26000, "2 Ton"),
    (26001, 30000, "2.5 Ton"),
    (30001, 38000, "3 Ton"),
    (38001, 44000, "3.5 Ton"),
    (44001, 50000, "4 Ton"),
    (50001, 56000, "4.5 Ton"),
    (56001, 999999, "5 Ton"),
]

# ─── AC Families (필터용) ───
AC_FAMILIES = {
    "AIR CONDITIONER",
    "MINI SPLIT AIR CONDITIONER",
    "WINDOW AIR CONDITIONER",
    "FREE STANDING AIR CONDITIONER",
    "PORTABLE",
    "AIR CURTAINS",
    "SEEC WINDOW AIR CONDITIONER",
}

# ─── Non-SA countries (제외) ───
EXCLUDE_COUNTRIES = {"BH", "OM", "BAHRAIN", "OMAN"}

# ─── Size Normalization Map ───
SIZE_NORMALIZE = {
    "1 ton": "1 Ton", "1.0 Ton": "1 Ton", "1.0 ton": "1 Ton",
    "1.5 ton": "1.5 Ton", "1.5 TON": "1.5 Ton",
    "2 ton": "2 Ton", "2.0 Ton": "2 Ton", "2.0 ton": "2 Ton", "2 TON": "2 Ton",
    "2.5 ton": "2.5 Ton", "2.5 TON": "2.5 Ton",
    "3 ton": "3 Ton", "3.0 Ton": "3 Ton",
    "3.5 ton": "3.5 Ton",
    "4 ton": "4 Ton", "4.0 Ton": "4 Ton",
    "4.5 ton": "4.5 Ton",
    "5 ton": "5 Ton", "5.0 Ton": "5 Ton",
}


def normalize_size(size):
    """Size 문자열 정규화 (예: '2 ton' → '2 Ton')"""
    if not size:
        return "1.5 Ton"
    s = str(size).strip()
    if s in SIZE_NORMALIZE:
        return SIZE_NORMALIZE[s]
    # Try regex normalization
    m = re.match(r"(\d+\.?\d*)\s*[Tt][Oo][Nn]", s)
    if m:
        val = float(m.group(1))
        if val == int(val):
            return f"{int(val)} Ton"
        return f"{val:g} Ton"
    return s


def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)


def btu_to_ton(btu_value):
    """BTU 값을 Ton으로 변환"""
    for low, high, ton in BTU_TO_TON:
        if low <= btu_value <= high:
            return ton
    return None


def extract_type_from_desc(desc):
    """Description에서 TYPE 추출 (Cold/Hot, Inverter/Rotary)"""
    desc_lower = desc.lower()

    # Inverter vs Rotary
    has_inverter = "inverter" in desc_lower or "inv " in desc_lower
    has_rotary = "rotary" in desc_lower

    # Hot vs Cold
    has_hot = any(kw in desc_lower for kw in [
        "hot", "heat", "h&c", "heat & cold", "hot and cold",
        "heat&cool", "heat/cold", "hot/cold"
    ])

    # Determine compressor type
    if has_inverter:
        comp = "Inverter"
    elif has_rotary:
        comp = "Rotary"
    else:
        comp = "Rotary"  # Default (97% accuracy)

    # Determine temperature
    if has_hot:
        temp = "Hot And Cold"
    else:
        temp = "Cold"

    return f"{temp} - {comp}"


def extract_size_from_desc(desc):
    """Description에서 SIZE 추출 (BTU → Ton 변환)"""
    # Try BTU pattern first
    btu_matches = re.findall(r"(\d{4,6})\s*(?:BTU|btu|Btu|btu/h)", desc)
    if btu_matches:
        btu_val = int(btu_matches[0])
        ton = btu_to_ton(btu_val)
        if ton:
            return ton

    # Try direct Ton pattern (e.g., "1.5T", "2 Ton", "1.5 TON")
    ton_match = re.search(r"(\d+\.?\d*)\s*(?:Ton|TON|T\b)", desc)
    if ton_match:
        ton_val = float(ton_match.group(1))
        ton_str = f"{ton_val:g} Ton"
        return ton_str

    return None


def extract_subfam_from_desc(desc):
    """Description에서 Sub Family 추출"""
    desc_lower = desc.lower()
    if "window" in desc_lower:
        if "seec" in desc_lower:
            return "SEEC WINDOW AIR CONDITIONER"
        return "WINDOW AIR CONDITIONER"
    elif "split" in desc_lower:
        return "MINI SPLIT AIR CONDITIONER"
    elif "portable" in desc_lower:
        return "PORTABLE"
    elif "free standing" in desc_lower or "floor standing" in desc_lower:
        return "FREE STANDING AIR CONDITIONER"
    elif "air curtain" in desc_lower:
        return "AIR CURTAINS"
    return "MINI SPLIT AIR CONDITIONER"  # Default


def get_extra_week_number(date_or_str):
    """eXtra 방식 주차 계산: 일요일 시작, 연도 경계에서 W52로 끊기

    규칙:
    - 주는 일요일에 시작하고 토요일에 끝남
    - 날짜의 실제 연도(year) 기준으로 주차 계산 (연도 경계에서 끊기)
    - 최대 W52 (W53 없음) — 12월 말이 W53이 될 때 W52로 캡
    - 예: 2026년 W1 = Jan 1(Thu)~Jan 3(Sat), W2 = Jan 4(Sun)~Jan 10(Sat)
    """
    if isinstance(date_or_str, datetime):
        dt = date_or_str
    else:
        dt = datetime.strptime(str(date_or_str).split()[0], "%Y-%m-%d")

    year = dt.year
    jan1 = datetime(year, 1, 1)

    # Jan 1의 주 시작 일요일 (직전 또는 당일 일요일)
    # weekday(): Mon=0, ..., Sun=6  →  일요일로 되돌아갈 일수: Sun=0, Mon=1, ..., Sat=6
    jan1_days_back = (jan1.weekday() + 1) % 7
    year_week1_start = jan1 - timedelta(days=jan1_days_back)

    # dt의 주 시작 일요일
    dt_days_back = (dt.weekday() + 1) % 7
    dt_week_start = dt - timedelta(days=dt_days_back)

    # 주차 = (dt_week_start - year_week1_start) / 7 + 1
    week_num = (dt_week_start - year_week1_start).days // 7 + 1

    # W52 캡 (12월 말 W53 방지)
    if week_num > 52:
        week_num = 52
    if week_num < 1:
        week_num = 1

    return week_num


def get_week_number(date_str):
    """날짜에서 eXtra 주차 번호 반환 (일요일 시작, W52 캡)"""
    return get_extra_week_number(date_str)


def get_day_key(date_str):
    """날짜에서 MM-DD 형식 반환"""
    if isinstance(date_str, datetime):
        dt = date_str
    else:
        dt = datetime.strptime(str(date_str).split()[0], "%Y-%m-%d")
    return dt.strftime("%m-%d")


def is_ac_family(family_desc):
    """AC 카테고리 여부 확인"""
    if not family_desc:
        return False
    fam = family_desc.strip().upper()
    return fam in AC_FAMILIES or "AIR CONDITIONER" in fam


def is_excluded_org(org_name):
    """비SA 매장 제외"""
    if not org_name:
        return True
    org = str(org_name).upper()
    return any(country in org for country in EXCLUDE_COUNTRIES)


class SelloutDataGenerator:
    def __init__(self):
        self.item_master = load_json(ITEM_MASTER_PATH)
        self.promoter_config = load_json(PROMOTER_CONFIG_PATH)
        self.branch_region = load_json(BRANCH_REGION_PATH)
        self.promoter_stores = set(self.promoter_config.get("promoter_stores", []))

        # Dimension indices
        self.years = [2024, 2025, 2026]
        self.days = [f"{m:02d}-{d:02d}" for m in range(1, 13) for d in range(1, 32)
                     if not (m in [4, 6, 9, 11] and d > 30) and not (m == 2 and d > 29)]
        self.weeks = [f"W{i}" for i in range(1, 53)]
        self.brands = set()
        self.subfamilies = set()
        self.types = {"Cold - Inverter", "Cold - Rotary",
                      "Hot And Cold - Inverter", "Hot And Cold - Rotary"}
        self.sizes = set()
        self.regions = set()
        self.branches = set()
        self.promoters = ["O", "X"]

        # Records
        self.records = []
        self.unclassified = []
        self.new_master_entries = {}

    def load_existing_data(self):
        """기존 data.json에서 2024-2025 데이터 로드"""
        if not os.path.exists(DATA_JSON_PATH):
            return

        print("Loading existing data.json for 2024-2025...")
        data = load_json(DATA_JSON_PATH)
        d = data["d"]

        # Load dimension values
        # Don't load dimensions from old data directly - collect from records
        # to ensure normalization is applied

        # Load 2024-2025 records only
        for rec in data["c"]:
            yi = rec[0]
            year = d["y"][yi]
            if year >= 2026:
                continue  # Skip 2026 - will be rebuilt

            day = d["day"][rec[1]]
            brand = d["b"][rec[2]]
            sf = d["sf"][rec[3]]
            type_ = d["t"][rec[4]]
            size = normalize_size(d["sz"][rec[5]])
            region = d["r"][rec[6]]
            prom = d["p"][rec[7]]
            branch = d["br"][rec[8]]
            qty = rec[9]
            val = rec[10]
            fp = rec[11]

            self.brands.add(brand)
            self.subfamilies.add(sf)
            self.sizes.add(size)
            self.regions.add(region)
            self.branches.add(branch)

            self.records.append({
                "year": year, "day": day, "brand": brand,
                "sf": sf, "type": type_, "size": size,
                "region": region, "promoter": prom, "branch": branch,
                "qty": qty, "val": val, "fp": fp
            })

        print(f"  Loaded {len(self.records)} records (2024-2025)")

    def load_annual_xlsx(self, path, year):
        """연간 sellout xlsx 로드 (2024 or 2025)"""
        import openpyxl
        print(f"Loading {os.path.basename(path)} (year={year})...")
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Year, ORG, Month, ITEM_NUM, ITEM_DESC, VENDOR, CAT, CAT_DESC, FAM, SUB_FAM, BRAND, TYPE, SIZE, Origin, QTY, SALE_VAL, QTY_RET, NET_SALES, Unit_price
            org = str(row[1]) if row[1] else ""
            sub_fam = str(row[9]) if row[9] else ""
            brand = str(row[10]) if row[10] else ""
            type_ = str(row[11]) if row[11] else ""
            size = str(row[12]) if row[12] else ""
            qty = row[14] if row[14] else 0
            sale_val = row[15] if row[15] else 0
            month = row[2] if row[2] else 1

            # Derive day from month (use 15th as default for monthly data)
            day_key = f"{int(month):02d}-15"

            # Region & Promoter
            region = self.branch_region.get(org, "Central")
            promoter = "O" if org in self.promoter_stores else "X"

            # Val & FP
            val = round(sale_val) if sale_val else 0
            fp = round(val * 1.15) if val else 0

            self.brands.add(brand)
            self.subfamilies.add(sub_fam)
            self.sizes.add(size)
            self.regions.add(region)
            self.branches.add(org)

            size = normalize_size(size)

            self.records.append({
                "year": year, "day": day_key, "brand": brand,
                "sf": sub_fam, "type": type_, "size": size,
                "region": region, "promoter": promoter, "branch": org,
                "qty": int(qty) if qty else 0, "val": val, "fp": fp
            })
            count += 1

        wb.close()
        print(f"  Loaded {count} records")

    def process_weekly_xlsx(self, week_num):
        """주간 xlsx 파일 처리 (2026)"""
        import openpyxl
        fname = f"week{week_num:02d}.xlsx"
        fpath = os.path.join(WEEKLY_DIR, fname)
        if not os.path.exists(fpath):
            return 0

        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]

        # Detect column layout
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(h).strip() if h else "" for h in row]

        # Find column indices
        col_map = {}
        for i, h in enumerate(headers):
            h_upper = h.upper()
            if "COUNTRY" in h_upper:
                col_map["country"] = i
            elif "CALENDAR" in h_upper or "DATE" in h_upper:
                col_map["date"] = i
            elif "ORGANIZATION" in h_upper:
                col_map["org"] = i
            elif "ITEM NUMBER" in h_upper or h_upper == "ITEM_NUMBER":
                col_map["item_num"] = i
            elif "ITEM MODEL" in h_upper:
                col_map["item_model"] = i
            elif "ITEM DESCRIPTION" in h_upper or "ITEM_DESCRIPTION" in h_upper:
                col_map["item_desc"] = i
            elif "VENDOR" in h_upper:
                col_map["vendor"] = i
            elif "SUB FAMILY" in h_upper or "SUB_FAMILY" in h_upper:
                col_map["subfam"] = i
            elif "FAMILY DESCRIPTION" in h_upper or "FAMILY_DESCRIPTION" in h_upper:
                col_map["family"] = i
            elif "BRAND" in h_upper:
                col_map["brand"] = i
            elif "SALE QUANTITY" in h_upper or "SALE_QUANTITY" in h_upper:
                col_map["qty"] = i
            elif "SALE VALUE" in h_upper or "SALE_VALUE" in h_upper:
                col_map["val"] = i

        # Fallback for standard 13-column format
        if "date" not in col_map:
            col_map = {
                "country": 0, "date": 1, "org": 2, "item_num": 3,
                "item_model": 4, "item_desc": 5, "vendor": 6,
                "cat_desc": 7, "family": 8, "subfam": 9,
                "brand": 10, "qty": 11, "val": 12
            }

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = list(row)

            # Get values
            org = str(vals[col_map["org"]]) if vals[col_map["org"]] else ""
            family = str(vals[col_map["family"]]) if vals[col_map.get("family", -1)] else ""
            item_num = str(vals[col_map["item_num"]]) if vals[col_map["item_num"]] else ""
            item_desc = str(vals[col_map.get("item_desc", col_map.get("item_model", 5))]) if len(vals) > 5 else ""
            brand = str(vals[col_map["brand"]]) if vals[col_map["brand"]] else ""
            subfam = str(vals[col_map.get("subfam", 9)]) if vals[col_map.get("subfam", 9)] else ""
            qty_raw = vals[col_map["qty"]]
            val_raw = vals[col_map["val"]]
            date_raw = vals[col_map["date"]]

            # Skip non-SA: Country 컬럼 우선 사용, 없으면 org name fallback
            country_val = str(vals[col_map["country"]]).strip().upper() if col_map.get("country") is not None and vals[col_map["country"]] else ""
            if country_val:
                if country_val != "SA":
                    continue
            else:
                if is_excluded_org(org):
                    continue

            # Filter AC only
            if not is_ac_family(family) and not is_ac_family(subfam):
                # Check item master
                if item_num in self.item_master:
                    pass  # Known AC item
                else:
                    continue  # Not AC

            # Parse date
            if isinstance(date_raw, datetime):
                dt = date_raw
            elif date_raw:
                try:
                    dt = datetime.strptime(str(date_raw).split()[0], "%Y-%m-%d")
                except ValueError:
                    continue
            else:
                continue

            day_key = dt.strftime("%m-%d")
            actual_year = dt.year  # Use actual year from date (e.g., week01 may have Dec 2025 dates)

            # Resolve TYPE & SIZE from master or description
            if item_num in self.item_master:
                master = self.item_master[item_num]
                type_ = master["type"]
                size = master["size"]
                if not subfam or subfam == "None":
                    subfam = master.get("sub_family", "MINI SPLIT AIR CONDITIONER")
            else:
                # New item - extract from description
                type_ = extract_type_from_desc(item_desc)
                size = extract_size_from_desc(item_desc)
                if not subfam or subfam == "None":
                    subfam = extract_subfam_from_desc(item_desc)

                if not size:
                    # Unclassified
                    self.unclassified.append({
                        "item_num": item_num,
                        "item_desc": item_desc,
                        "brand": brand,
                        "missing": "size",
                        "week": week_num
                    })
                    size = "1.5 Ton"  # Default

                # Add to new master entries
                if item_num not in self.new_master_entries:
                    self.new_master_entries[item_num] = {
                        "description": item_desc,
                        "sub_family": subfam,
                        "brand": brand,
                        "type": type_,
                        "size": size,
                        "auto_extracted": True
                    }

            # Normalize type
            if type_ and type_ not in self.types:
                # Try to normalize (e.g., "Hot and Cold - Rotary" → "Hot And Cold - Rotary")
                type_norm = type_.replace("and", "And").replace(" - ", " - ")
                if type_norm in self.types:
                    type_ = type_norm
                else:
                    type_ = "Cold - Rotary"  # Default

            # Region
            region = self.branch_region.get(org, "")
            if not region:
                # Try partial match
                for known_br, known_reg in self.branch_region.items():
                    if org.startswith(known_br[:5]) or known_br.startswith(org[:5]):
                        region = known_reg
                        break
                if not region:
                    region = "Central"  # Default

            # Promoter
            promoter = "O" if org in self.promoter_stores else "X"

            # Normalize size
            size = normalize_size(size)

            # Values
            qty = int(qty_raw) if qty_raw else 0
            val = round(float(val_raw)) if val_raw else 0
            fp = round(val * 1.15)

            self.brands.add(brand)
            self.subfamilies.add(subfam)
            self.sizes.add(size)
            self.regions.add(region)
            self.branches.add(org)

            self.records.append({
                "year": actual_year, "day": day_key, "brand": brand,
                "sf": subfam, "type": type_, "size": size,
                "region": region, "promoter": promoter, "branch": org,
                "qty": qty, "val": val, "fp": fp
            })
            count += 1

        wb.close()
        return count

    def build_data_json(self):
        """data.json 형식으로 변환"""
        # Sort dimensions
        brands_list = sorted(self.brands - {""})
        sf_list = sorted(self.subfamilies - {""})
        types_list = sorted(self.types)
        sizes_list = sorted(self.sizes - {""}, key=lambda s: (
            float(re.search(r"(\d+\.?\d*)", s).group(1)) if re.search(r"(\d+\.?\d*)", s) else 999
        ))
        regions_list = sorted(self.regions - {""})
        branches_list = sorted(self.branches - {""})

        # Build index maps
        year_idx = {y: i for i, y in enumerate(self.years)}
        day_idx = {d: i for i, d in enumerate(self.days)}
        brand_idx = {b: i for i, b in enumerate(brands_list)}
        sf_idx = {s: i for i, s in enumerate(sf_list)}
        type_idx = {t: i for i, t in enumerate(types_list)}
        size_idx = {s: i for i, s in enumerate(sizes_list)}
        region_idx = {r: i for i, r in enumerate(regions_list)}
        prom_idx = {"O": 0, "X": 1}
        branch_idx = {b: i for i, b in enumerate(branches_list)}

        # Build type groups
        tg = {}
        for i, t in enumerate(types_list):
            tg[str(i)] = "Inverter" if "Inverter" in t else "Rotary"

        # Build date metadata
        dm = {}
        for yi, year in enumerate(self.years):
            for di, day in enumerate(self.days):
                try:
                    # Handle leap year
                    month = int(day.split("-")[0])
                    day_num = int(day.split("-")[1])
                    if month == 2 and day_num == 29 and year % 4 != 0:
                        continue
                    dt = datetime(year, month, day_num)
                    week = get_extra_week_number(dt)
                    quarter = (month - 1) // 3 + 1
                    dm[f"{yi}-{di}"] = [week - 1, month, quarter]
                except ValueError:
                    continue

        # Convert records to compressed format
        c_records = []
        skipped = 0
        for rec in self.records:
            try:
                yi = year_idx.get(rec["year"])
                di = day_idx.get(rec["day"])
                bi = brand_idx.get(rec["brand"])
                sfi = sf_idx.get(rec["sf"])
                ti = type_idx.get(rec["type"])
                szi = size_idx.get(rec["size"])
                ri = region_idx.get(rec["region"])
                pi = prom_idx.get(rec["promoter"], 1)
                bri = branch_idx.get(rec["branch"])

                if any(v is None for v in [yi, di, bi, sfi, ti, szi, ri, bri]):
                    skipped += 1
                    continue

                c_records.append([
                    yi, di, bi, sfi, ti, szi, ri, pi, bri,
                    rec["qty"], rec["val"], rec["fp"]
                ])
            except (KeyError, TypeError):
                skipped += 1

        if skipped:
            print(f"  Skipped {skipped} records (missing dimensions)")

        data = {
            "d": {
                "y": self.years,
                "day": self.days,
                "w": self.weeks,
                "b": brands_list,
                "sf": sf_list,
                "t": types_list,
                "sz": sizes_list,
                "r": regions_list,
                "p": self.promoters,
                "br": branches_list
            },
            "tg": tg,
            "dm": dm,
            "c": c_records
        }

        return data

    def update_item_master(self):
        """신규 아이템을 마스터에 추가"""
        if not self.new_master_entries:
            return

        for item_num, info in self.new_master_entries.items():
            if item_num not in self.item_master:
                self.item_master[item_num] = info

        save_json(ITEM_MASTER_PATH, self.item_master)
        print(f"  Updated item master: +{len(self.new_master_entries)} new items")

    def save_unclassified(self):
        """미분류 아이템 리포트 저장"""
        if not self.unclassified:
            return
        save_json(UNCLASSIFIED_PATH, self.unclassified)
        print(f"  Unclassified items report: {len(self.unclassified)} items → {UNCLASSIFIED_PATH}")

    def run(self, rebuild=False):
        """메인 실행"""
        print("=" * 60)
        print("eXtra Sellout Dashboard - Data Generator")
        print("=" * 60)

        if rebuild:
            print("\n[MODE] Full rebuild (2024 + 2025 + 2026)")
            # Load 2024 & 2025 from xlsx
            xlsx_2024 = os.path.join(FCST_DIR, "extra_2024.xlsx")
            xlsx_2025 = os.path.join(FCST_DIR, "extra_2025.xlsx")
            if os.path.exists(xlsx_2024):
                self.load_annual_xlsx(xlsx_2024, 2024)
            if os.path.exists(xlsx_2025):
                self.load_annual_xlsx(xlsx_2025, 2025)
        else:
            print("\n[MODE] Incremental update (keep 2024-2025, rebuild 2026)")
            self.load_existing_data()

        # Process 2026 weekly files
        print("\nProcessing 2026 weekly xlsx files...")
        total_weekly = 0
        for week_num in range(1, 53):
            count = self.process_weekly_xlsx(week_num)
            if count > 0:
                print(f"  week{week_num:02d}.xlsx: {count} AC records")
                total_weekly += count

        print(f"\nTotal 2026 weekly records: {total_weekly}")
        print(f"Total all records: {len(self.records)}")

        # Build data.json
        print("\nBuilding data.json...")
        data = self.build_data_json()
        print(f"  Records in output: {len(data['c'])}")
        print(f"  Brands: {len(data['d']['b'])}")
        print(f"  Branches: {len(data['d']['br'])}")
        print(f"  Sizes: {data['d']['sz']}")

        # Save
        save_json(DATA_JSON_PATH, data)
        file_size = os.path.getsize(DATA_JSON_PATH) / 1024 / 1024
        print(f"\n  Saved: {DATA_JSON_PATH} ({file_size:.1f} MB)")

        # Update master & save unclassified
        self.update_item_master()
        self.save_unclassified()

        # Summary
        print("\n" + "=" * 60)
        print("SUMMARY")
        print("=" * 60)
        rec_by_year = defaultdict(int)
        for rec in self.records:
            rec_by_year[rec["year"]] += 1
        for year in sorted(rec_by_year):
            print(f"  {year}: {rec_by_year[year]:,} records")
        print(f"  Total: {len(data['c']):,} records")
        if self.new_master_entries:
            print(f"  New items added to master: {len(self.new_master_entries)}")
        if self.unclassified:
            print(f"  Unclassified items: {len(self.unclassified)}")
        print("=" * 60)

        return data


if __name__ == "__main__":
    rebuild = "--rebuild" in sys.argv
    generator = SelloutDataGenerator()
    generator.run(rebuild=rebuild)
